import os
from datetime import date, datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.mail import send_mail
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Booking
from .forms import BookingForm, RegistroForm
from catalog.models import Product


# ----------------- PÁGINAS PÚBLICAS -----------------

def Inicio(request):
    return render(request, 'Inicio.html')


def Catalogo(request):
    return redirect("catalog:lista")


# ----------------- DASHBOARD (Solo Admin) -----------------

def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


@login_required
@user_passes_test(is_admin)
def dashboard(request):
    return render(request, "dashboard.html")


# ----------------- RESERVAR (Solo usuarios logueados) -----------------

@login_required
def Reservar(request):
    carrito = request.session.get('carrito', [])
    productos = Product.objects.filter(pk__in=carrito)
    total_productos = sum(p.price for p in productos)
    precio_por_hora = 20000

    if request.method == "POST":
        form = BookingForm(request.POST)
        if form.is_valid():
            booking = form.save(commit=False)
            booking.user = request.user

            # 🚫 Evita solapamientos
            conflicto = Booking.objects.filter(
                event_date=booking.event_date,
                start_time__lt=booking.end_time,
                end_time__gt=booking.start_time,
            ).exclude(status='cancelled')

            if conflicto.exists():
                form.add_error(None, "⛔ Ese horario ya está reservado. Elige otro.")
                return render(request, "Reservar.html", {
                    "form": form,
                    "productos": productos,
                    "total_productos": total_productos,
                    "precio_por_hora": precio_por_hora,
                })

            # 💰 Calcular total
            start = datetime.combine(date.today(), booking.start_time)
            end = datetime.combine(date.today(), booking.end_time)
            horas = (end - start).seconds // 3600
            booking.total = total_productos + (horas * precio_por_hora)
            booking.status = 'pending'
            booking.save()
            booking.products.set(productos)

            # Limpia el carrito
            request.session['carrito'] = []
            messages.success(request, "✅ ¡Reserva realizada exitosamente!")
            return redirect("Reservas")

        messages.error(request, "Revisa los errores del formulario.")
    else:
        initial = {
            "event_date": request.GET.get("date"),
            "start_time": request.GET.get("start"),
            "end_time": request.GET.get("end"),
        }
        initial = {k: v for k, v in initial.items() if v}
        form = BookingForm(initial=initial)

    return render(request, "Reservar.html", {
        "form": form,
        "productos": productos,
        "total_productos": total_productos,
        "precio_por_hora": precio_por_hora,
    })


# ----------------- EDITAR RESERVA (Solo Admin) -----------------

@login_required
@user_passes_test(is_admin)
def editar_reserva(request, pk):
    """
    Permite al admin editar una reserva sin perder la fecha o la hora original.
    Corrige el caso en que el campo de fecha no se envía en el POST.
    """
    reserva = get_object_or_404(Booking, pk=pk)

    if request.method == "POST":
        # ✅ Si no se envía la fecha, reinyectamos la actual antes de validar
        post_data = request.POST.copy()
        if not post_data.get("event_date") and reserva.event_date:
            post_data["event_date"] = reserva.event_date.strftime("%Y-%m-%d")

        form = BookingForm(post_data, instance=reserva)

        if form.is_valid():
            updated = form.save(commit=False)

            # 🕐 Convertir hora a formato correcto (por seguridad)
            try:
                if isinstance(updated.start_time, str):
                    updated.start_time = datetime.strptime(updated.start_time, "%H:%M").time()
                if isinstance(updated.end_time, str):
                    updated.end_time = datetime.strptime(updated.end_time, "%H:%M").time()
            except Exception:
                pass

            updated.save()
            form.save_m2m()
            messages.success(request, "✅ Reserva actualizada correctamente.")
            return redirect("catalog:calendario")
        else:
            messages.error(request, "⚠️ Revisa los campos, hay errores en el formulario.")
    else:
        form = BookingForm(instance=reserva)

    return render(request, "booking/editar.html", {"form": form, "reserva": reserva})


# ----------------- LISTA RESERVAS -----------------

@login_required
def Reservas(request):
    if request.user.is_staff or request.user.is_superuser:
        qs = Booking.objects.exclude(status='cancelled')
    else:
        qs = Booking.objects.filter(user=request.user).exclude(status='cancelled')

    qs = qs.order_by("-event_date", "-start_time")
    paginator = Paginator(qs, 20)
    reservas = paginator.get_page(request.GET.get("page"))
    return render(request, "Reservas.html", {"reservas": reservas})


# ----------------- ACTUALIZAR ESTADO (Solo Admin) -----------------

@login_required
@user_passes_test(is_admin)
def actualizar_estado(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    if request.method == "POST":
        nuevo_estado = request.POST.get("estado")
        booking.status = nuevo_estado
        booking.save()

        # 📧 Notificación por correo
        to_email = (booking.customer_email or "").strip()
        if to_email:
            asunto, mensaje = None, None
            datos_reserva = (
                f"Fecha: {booking.event_date:%d-%m-%Y}\n"
                f"Horario: {booking.start_time.strftime('%H:%M')} - {booking.end_time.strftime('%H:%M')}\n"
                f"Lugar: {booking.location or '—'}\n"
                f"Total: ${booking.total:,}".replace(",", ".")
            )

            if nuevo_estado == "approved":
                asunto = "✅ ¡Tu reserva fue aprobada!"
                mensaje = (
                    f"Hola {booking.customer_name},\n\n"
                    f"Tu reserva ha sido APROBADA.\n\n{datos_reserva}\n\n"
                    f"Gracias por confiar en HappyHuapi.\n— Equipo HappyHuapi"
                )
            elif nuevo_estado == "cancelled":
                asunto = "❌ Tu reserva fue cancelada"
                mensaje = (
                    f"Hola {booking.customer_name},\n\n"
                    f"Lamentamos informarte que tu reserva fue CANCELADA.\n\n{datos_reserva}\n\n"
                    f"Si crees que es un error o deseas reprogramar, contáctanos.\n— Equipo HappyHuapi"
                )
            elif nuevo_estado == "done":
                asunto = "🟦 Tu reserva fue marcada como realizada"
                mensaje = (
                    f"Hola {booking.customer_name},\n\n"
                    f"Tu reserva fue marcada como REALIZADA.\n\n{datos_reserva}\n\n"
                    f"Gracias por preferirnos.\n— Equipo HappyHuapi"
                )

            if asunto and mensaje:
                try:
                    send_mail(
                        subject=asunto,
                        message=mensaje,
                        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                        recipient_list=[to_email],
                        fail_silently=True,
                    )
                except Exception:
                    pass

        messages.success(request, "Estado actualizado correctamente.")
        return redirect("Reservas")

    messages.error(request, "Solicitud inválida.")
    return redirect("Reservas")


# ----------------- LOGOUT -----------------

def logout_view(request):
    logout(request)
    return redirect('Inicio')


# ----------------- REGISTRO (CLIENTE) -----------------

def registro(request):
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get("email")
            if email and form._meta.model.objects.filter(email=email).exists():
                messages.error(request, "⚠️ Ya existe una cuenta con este correo.")
                return render(request, "auth/registro.html", {"form": form})
            user = form.save()
            login(request, user)
            messages.success(request, "✅ Cuenta creada correctamente. ¡Bienvenido!")
            return redirect("Inicio")
    else:
        form = RegistroForm()
    return render(request, "auth/registro.html", {"form": form})


# ----------------- EXPORTAR RESERVAS A EXCEL (SOLO ADMIN) -----------------

@login_required
@user_passes_test(is_admin)
def exportar_reservas(request):
    reservas = Booking.objects.exclude(status='cancelled').order_by("-event_date", "-start_time")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas"

    # Portada
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    titulo = ws["A1"]
    titulo.value = "📄 Reporte de Reservas - HappyHuapi"
    titulo.font = Font(size=18, bold=True, color="1F4E78")
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    fecha = ws["A2"]
    fecha.value = f"Generado el: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    fecha.font = Font(size=12, italic=True, color="404040")
    fecha.alignment = Alignment(horizontal="center", vertical="center")

    ws.append(["", ""])
    columnas = ["Fecha", "Hora Inicio", "Hora Término", "Cliente",
                "Correo", "Teléfono", "Lugar", "Total", "Estado", "Productos"]
    ws.append(columnas)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for col in range(1, len(columnas) + 1):
        c = ws.cell(row=5, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    for r in reservas:
        productos = ", ".join([p.name for p in r.products.all()]) or "Sin productos"
        ws.append([
            r.event_date.strftime("%d-%m-%Y"),
            r.start_time.strftime("%H:%M"),
            r.end_time.strftime("%H:%M"),
            r.customer_name, r.customer_email, r.customer_phone,
            r.location or "-", f"${r.total:,}".replace(",", "."),
            r.get_status_display(), productos
        ])

    for row in ws.iter_rows(min_row=6):
        for cell in row:
            cell.border = thin
            cell.alignment = center

    widths = [12, 12, 12, 25, 25, 14, 25, 12, 14, 50]
    for col, width in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename=\"Reservas_HappyHuapi.xlsx\"'
    wb.save(response)
    return response
