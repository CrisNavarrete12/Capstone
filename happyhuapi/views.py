import os 
from datetime import date, datetime
from decimal import Decimal
import uuid
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test

from catalog.models import Product
from .models import Booking

from django.utils import timezone
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.mail import EmailMultiAlternatives, send_mail

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from transbank.webpay.webpay_plus.transaction import Transaction
from transbank.common.integration_type import IntegrationType

from .models import Booking
from .forms import BookingForm, RegistroForm
from catalog.models import Product

# QR UTILS
from happyhuapi.utils_qr import generar_qr_para_reserva


#   PERMISOS PARA ADMIN Y TRABAJADOR

def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

def is_worker(user):
    return user.groups.filter(name="Trabajador").exists()

def admin_o_trabajador(user):
    return is_admin(user) or is_worker(user)


#   CONFIG WEBPAY PRUEBA
Transaction.commerce_code = settings.WEBPAY_COMMERCE_CODE
Transaction.api_key = settings.WEBPAY_API_KEY
Transaction.integration_type = "TEST"


# ----------------- PÁGINAS PÚBLICAS -----------------

def Inicio(request):
    return render(request, "Inicio.html")


def Catalogo(request):
    return redirect("catalog:lista")


def iniciar_pago(request):
    return render(request, "pago_pendiente.html")


# ----------------- RESERVAR (Pago 30% WebPay) -----------------

@login_required
def Reservar(request):
    carrito = request.session.get("carrito", [])
    productos = Product.objects.filter(pk__in=carrito)
    total_productos = sum(p.price for p in productos)

    precio_por_hora = 20000

    if request.method == "POST":
        form = BookingForm(request.POST)

        if form.is_valid():
            booking_temp = form.save(commit=False)

            start = datetime.combine(date.today(), booking_temp.start_time)
            end = datetime.combine(date.today(), booking_temp.end_time)
            horas = (end - start).seconds // 3600

            total_final = total_productos + (horas * precio_por_hora)
            anticipo = int(total_final * 0.30)
            
            if anticipo <= 0:
                messages.error(request, "El monto del pago debe ser mayor a 0. Agrega productos y horas.")
                return redirect("Reservar")

            request.session["booking_data"] = {
                "customer_name": booking_temp.customer_name,
                "customer_email": booking_temp.customer_email,
                "customer_phone": booking_temp.customer_phone,
                "event_date": booking_temp.event_date.strftime("%Y-%m-%d"),
                "start_time": booking_temp.start_time.strftime("%H:%M"),
                "end_time": booking_temp.end_time.strftime("%H:%M"),
                "location": booking_temp.location,
                "notes": booking_temp.notes,
                "total_price": int(total_final),
                "deposit_amount": int(anticipo),
            }

            buy_order = str(uuid.uuid4())[:26]
            session_id = "session123"
            amount = anticipo

            tx = Transaction().create(
                buy_order=buy_order,
                session_id=session_id,
                amount=amount,
                return_url=f"{settings.BASE_URL}/pago/exito/",
            )

            return redirect(f'{tx["url"]}?token_ws={tx["token"]}')

        messages.error(request, "Revisa los errores del formulario.")

    else:
        initial = {
            "event_date": request.GET.get("date"),
            "start_time": request.GET.get("start"),
            "end_time": request.GET.get("end"),
        }
        initial = {k: v for k, v in initial.items() if v}
        form = BookingForm(initial=initial)

    return render(
        request,
        "Reservar.html",
        {
            "form": form,
            "productos": productos,
            "total_productos": total_productos,
            "precio_por_hora": precio_por_hora,
        },
    )
# ----------------- CONFIRMAR PAGO WEBPAY -----------------

@login_required
def pago_exito(request):
    token_ws = request.GET.get("token_ws") or request.POST.get("token_ws") or request.GET.get("token")

    if not token_ws:
        return render(request, "pago_error.html", {
            "mensaje": "No se recibió el token de pago."
        })

    #  VALIDAR ESTADO DEL PAGO
    try:
        response = Transaction().commit(token_ws)
        print("WEBPAY COMMIT RESPONSE:", response)
    except Exception as e:
        print(" ERROR EN COMMIT:", e)
        return render(request, "pago_error.html", {
            "mensaje": f"Error confirmando el pago: {e}"
        })
    #  Validar que exista información de reserva en la sesión
    booking_data = request.session.get("booking_data")

    if not booking_data:
        messages.error(request, "No se encontraron datos de la reserva.")
        return redirect("Reservar")

    estado = response.get("status")

    #  Si el pago NO fue autorizado → NO crear reserva
    if estado != "AUTHORIZED":
        return render(request, "pago_error.html", {
            "mensaje": f"El pago fue {estado.lower()} por el banco. La reserva no fue creada."
        })

    # ============================
    # ✔️ SI LLEGAMOS AQUÍ → PAGO REALMENTE EXITOSO
    # ============================

    booking_data = request.session.get("booking_data")
    carrito = request.session.get("carrito", [])

    if not booking_data:
        messages.error(request, "No se encontraron datos de la reserva.")
        return redirect("Reservar")
    

    # Convertir formatos
    event_date = datetime.strptime(booking_data["event_date"], "%Y-%m-%d").date()
    start_time = datetime.strptime(booking_data["start_time"], "%H:%M").time()
    end_time = datetime.strptime(booking_data["end_time"], "%H:%M").time()

    # Crear reserva final
    booking = Booking.objects.create(
        user=request.user,
        customer_name=booking_data["customer_name"],
        customer_email=booking_data["customer_email"],
        customer_phone=booking_data["customer_phone"],
        event_date=event_date,
        start_time=start_time,
        end_time=end_time,
        location=booking_data["location"],
        notes=booking_data["notes"],
        total_price=Decimal(booking_data["total_price"]),
        deposit_amount=Decimal(booking_data["deposit_amount"]),
        payment_date=timezone.now(),
        status="pending",
    )

    # Agregar productos
    productos_qs = Product.objects.filter(pk__in=carrito)
    booking.products.set(productos_qs)

    saldo_restante = booking.total_price - booking.deposit_amount

    # -------- GENERAR QR --------
    qr_path = generar_qr_para_reserva(booking)

    # -------- FORMATO PRECIOS --------
    total_str = f"{booking.total_price:,}".replace(",", ".")
    deposito_str = f"{booking.deposit_amount:,}".replace(",", ".")
    saldo_str = f"{saldo_restante:,}".replace(",", ".")

    # -------- PRODUCTOS --------
    productos_texto = "\n".join(
        [f"- {p.name}: ${str(f'{p.price:,}').replace(',', '.')}" for p in productos_qs]
    ) or "Sin productos seleccionados."

    productos_html = "".join(
        [f"<li>{p.name}: ${str(f'{p.price:,}').replace(',', '.')}</li>" for p in productos_qs]
    ) or "<li>Sin productos seleccionados.</li>"

    # -------- CORREO DE CONFIRMACIÓN --------
    mensaje_texto = f"""
Hola {booking.customer_name} 👋,

¡Tu pago del 30% fue recibido con éxito y tu reserva en HappyHuapi fue confirmada! 🎉

📅 Fecha del evento: {booking.event_date}
⏰ Horario: {booking.start_time} – {booking.end_time}
📍 Lugar: {booking.location}

------------------------
🛍️ Productos seleccionados:
{productos_texto}

------------------------
💰 Detalle de pago:
• Total: ${total_str}
• Pagado (30%): ${deposito_str}
• Pendiente: ${saldo_str}

Adjuntamos tu código QR para validar el evento.

¡Gracias por confiar en HappyHuapi! 💙
"""

    mensaje_html = f"""
<h2 style="color:#4A90E2;">🎉 ¡Reserva Confirmada, {booking.customer_name}! 🎉</h2>

<h3>📅 Detalles del evento</h3>
<ul>
  <li><strong>Fecha:</strong> {booking.event_date}</li>
  <li><strong>Horario:</strong> {booking.start_time} – {booking.end_time}</li>
  <li><strong>Lugar:</strong> {booking.location}</li>
</ul>

<h3>🛍️ Productos seleccionados</h3>
<ul>{productos_html}</ul>

<h3>💰 Resumen de pago</h3>
<ul>
  <li><strong>Total:</strong> ${total_str}</li>
  <li><strong>Pagado (30%):</strong> ${deposito_str}</li>
  <li><strong>Pendiente:</strong> ${saldo_str}</li>
</ul>

<p>Tu código QR viene adjunto en este correo.</p>
"""

    email = EmailMultiAlternatives(
        subject="🎉 Confirmación de reserva + QR - HappyHuapi",
        body=mensaje_texto,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[booking.customer_email],
    )
    email.attach_alternative(mensaje_html, "text/html")

    # Adjuntar QR
    try:
        full_path = os.path.join(settings.MEDIA_ROOT, booking.qr_image.name)
        with open(full_path, "rb") as qr_file:
            email.attach("codigo_qr.png", qr_file.read(), "image/png")
    except Exception as e:
        print("ERROR adjuntando QR:", e)

    email.send()

    # -------- LIMPIAR SESIÓN --------
    request.session["carrito"] = []
    request.session.pop("booking_data", None)
    request.session.pop("token_ws", None)
    request.session.modified = True

    # -------- RETORNO --------
    return render(
        request,
        "pago_exito.html",
        {
            "booking": booking,
            "saldo_restante": saldo_restante,
            "qr_path": qr_path,
        },
    )

# ----------------- RESERVAS -----------------

@login_required
def Reservas(request):
    if is_admin(request.user):
        qs = Booking.objects.exclude(status="cancelled")
    else:
        qs = Booking.objects.filter(user=request.user).exclude(status="cancelled")

    qs = qs.order_by("-event_date", "-start_time")
    paginator = Paginator(qs, 20)
    reservas = paginator.get_page(request.GET.get("page"))
    return render(request, "Reservas.html", {"reservas": reservas})


@login_required
@user_passes_test(is_admin)
def actualizar_estado(request, pk):
    booking = get_object_or_404(Booking, pk=pk)

    ESTADOS_VALIDOS = {
        "pending": ["approved", "cancelled"],
        "approved": ["done", "cancelled"],
        "done": [],
        "cancelled": [],
    }

    if request.method != "POST":
        messages.error(request, "Solicitud inválida.")
        return redirect("Reservas")

    nuevo_estado = request.POST.get("estado")
    estado_actual = booking.status

    # Validar transición
    if nuevo_estado not in ESTADOS_VALIDOS.get(estado_actual, []):
        messages.error(
            request,
            f"No puedes cambiar de '{estado_actual}' a '{nuevo_estado}'."
        )
        return redirect("Reservas")

    # Aplicar cambio
    booking.status = nuevo_estado
    booking.save()

    #     ENVÍO DE CORREO
    cliente = booking.customer_name
    correo = booking.customer_email
    fecha = booking.event_date
    hora = f"{booking.start_time} – {booking.end_time}"

    mensaje_html = ""
    asunto = "Actualización de tu reserva - HappyHuapi"

    if nuevo_estado == "approved":
        asunto = "Tu reserva ha sido aprobada ✔"
        mensaje_html = f"""
        <h2>Hola {cliente},</h2>
        <p>Tu reserva ha sido <b>APROBADA</b>.</p>
        <p><b>Fecha:</b> {fecha}<br>
        <b>Horario:</b> {hora}</p>
        <p>Nos vemos en tu evento 😊</p>
        """

    elif nuevo_estado == "done":
        asunto = "Tu evento ha sido marcado como realizado 🎉"
        mensaje_html = f"""
        <h2>Hola {cliente},</h2>
        <p>Marcamos tu evento como <b>REALIZADO</b>.</p>
        <p>¡Gracias por confiar en HappyHuapi!</p>
        """

    elif nuevo_estado == "cancelled":
        asunto = "Tu reserva ha sido cancelada ❌"
        mensaje_html = f"""
        <h2>Hola {cliente},</h2>
        <p>Tu reserva ha sido <b>CANCELADA</b>.</p>
        <p>Si esto es un error, contáctanos.</p>
        """

    if correo and mensaje_html:
        email = EmailMultiAlternatives(
            subject=asunto,
            body="Actualización de tu reserva.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[correo],
        )
        email.attach_alternative(mensaje_html, "text/html")
        email.send()

    messages.success(request, "Estado actualizado correctamente.")
    return redirect("Reservas")

# ----------------- LOGOUT -----------------

def logout_view(request):
    logout(request)
    return redirect("Inicio")


# ----------------- REGISTRO -----------------

from django.contrib.auth.models import Group

def registro(request):
    if request.method == "POST":
        form = RegistroForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get("email")

            # Validar email duplicado
            if email and form._meta.model.objects.filter(email=email).exists():
                messages.error(request, "Ya existe una cuenta con este correo.")
                return render(request, "auth/registro.html", {"form": form})

            # Crear usuario
            user = form.save()

            # AGREGAR AL GRUPO CLIENTE (la parte más importante)
            try:
                grupo = Group.objects.get(name="Cliente")
                user.groups.add(grupo)
            except Group.DoesNotExist:
                # Si el grupo no existe, se crea automáticamente
                grupo = Group.objects.create(name="Cliente")
                user.groups.add(grupo)

            user.save()

            # Iniciar sesión automáticamente
            login(request, user)

            messages.success(request, "Cuenta creada exitosamente.")
            return redirect("Inicio")

    else:
        form = RegistroForm()

    return render(request, "auth/registro.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def exportar_reservas(request):
    reservas = Booking.objects.exclude(status='cancelled').order_by("-event_date", "-start_time")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas"

    # Títulos principales
    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")

    titulo = ws["A1"]
    titulo.value = "Reporte de Reservas - HappyHuapi"
    titulo.font = Font(size=18, bold=True, color="1F4E78")
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    fecha = ws["A2"]
    fecha.value = f"Generado el: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    fecha.font = Font(size=12, italic=True, color="404040")
    fecha.alignment = Alignment(horizontal="center", vertical="center")

    # SIN FILA VACÍA AQUÍ

    columnas = [
        "Fecha", "Hora Inicio", "Hora Término", "Cliente",
        "Correo", "Teléfono", "Lugar",
        "Notas del Cliente",
        "Total", "Anticipo (30%)",
        "Estado", "Productos"
    ]

    ws.append(columnas)

    # Estilos del encabezado (fila 3)
    header_row = 3
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for col in range(1, len(columnas) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    # Altura visual correcta del header
    ws.row_dimensions[header_row].height = 25

    # Datos
    for r in reservas:
        productos = ", ".join([p.name for p in r.products.all()]) or "Sin productos"

        ws.append([
            r.event_date.strftime("%d-%m-%Y"),
            r.start_time.strftime("%H:%M"),
            r.end_time.strftime("%H:%M"),
            r.customer_name,
            r.customer_email,
            r.customer_phone,
            r.location or "-",
            r.notes or "—",
            f"${r.total_price:,}".replace(",", "."),
            f"${r.deposit_amount:,}".replace(",", "."),
            r.get_status_display(),
            productos
        ])

    # Estilo filas
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = thin
            cell.alignment = center

    # Ajuste de columnas
    widths = [12, 12, 12, 20, 25, 14, 20, 25, 12, 16, 14, 40]
    for col, width in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = width

    # Descargar archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Reservas_HappyHuapi.xlsx"'
    wb.save(response)
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.contrib import messages
from datetime import datetime

from .models import Booking
from catalog.models import Product
from .forms import BookingForm



#  EDITAR RESERVA  

@login_required
@user_passes_test(is_admin)
def editar_reserva(request, pk):
    reserva = get_object_or_404(Booking, pk=pk)

    if request.method == "POST":
        post_data = request.POST.copy()

        # Mantener fecha si viene vacía
        if not post_data.get("event_date") and reserva.event_date:
            post_data["event_date"] = reserva.event_date.strftime("%Y-%m-%d")

        form = BookingForm(post_data, instance=reserva)

        if form.is_valid():

            updated = form.save(commit=False)

            # USAR cleaned_data que YA VIENE CORREGIDO DEL FORM
            updated.start_time = form.cleaned_data["start_time"]
            updated.end_time = form.cleaned_data["end_time"]

            updated.save()
            form.save_m2m()

            messages.success(request, "Reserva actualizada.")
            return redirect("catalog:calendario")

        else:
            messages.error(request, "Errores en el formulario.")

    else:
        form = BookingForm(instance=reserva)

    return render(
        request,
        "booking/editar.html",
        {"form": form, "reserva": reserva},
    )



#  QUITAR PRODUCTO DE UNA RESERVA 

@login_required
@user_passes_test(is_admin)
def quitar_producto(request, pk, product_id):
    reserva = get_object_or_404(Booking, pk=pk)
    producto = get_object_or_404(Product, pk=product_id)

    # Eliminar la relación en la tabla intermedia
    reserva.products.remove(producto)

    return JsonResponse({"ok": True, "product_id": product_id})

# IMPORTS NECESARIOS

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.urls import reverse

from happyhuapi.models import Booking, Evidence, EvidencePhoto
from django.contrib.auth.models import Group

# FUNCIÓN — VALIDAR TRABAJADOR

def es_trabajador(user):
    return user.is_authenticated and user.groups.filter(name="Trabajador").exists()

# VISTA PRINCIPAL DEL QR (TRABAJADOR)

@login_required
def evidencia_qr(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    evidencia, created = Evidence.objects.get_or_create(booking=booking)

    if not es_trabajador(request.user):
        return render(request, "no_autorizado.html")

    # Si el trabajador envía una foto
    if request.method == "POST":
        image = request.FILES.get("image")

        if not image:
            return JsonResponse({"error": "No se envió imagen"}, status=400)

        foto = EvidencePhoto.objects.create(
            evidence=evidencia,
            image=image
        )

        return JsonResponse({
            "ok": True,
            "photo_url": foto.image.url
        })

    return render(request, "evidencia_trabajador.html", {
        "booking": booking,
        "evidencia": evidencia,
        "photos": evidencia.photos.all(),
    })


# REGISTRO DE LLEGADA

@require_POST
@login_required
def registrar_llegada(request, booking_id):
    evidencia = get_object_or_404(Evidence, booking_id=booking_id)

    if not es_trabajador(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)

    evidencia.arrival_time = timezone.now()
    evidencia.save()

    return JsonResponse({"ok": True, "arrival_time": evidencia.arrival_time})


# REGISTRO DE SALIDA

@require_POST
@login_required
def registrar_salida(request, booking_id):
    evidencia = get_object_or_404(Evidence, booking_id=booking_id)

    if not es_trabajador(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)

    evidencia.finish_time = timezone.now()
    evidencia.save()

    return JsonResponse({"ok": True, "finish_time": evidencia.finish_time})


# SUBIR FOTO DE EVIDENCIA

@require_POST
@login_required
def subir_foto_evidencia(request, booking_id):
    evidencia = get_object_or_404(Evidence, booking_id=booking_id)

    if not es_trabajador(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)

    image = request.FILES.get("image")
    if not image:
        return JsonResponse({"error": "No se envió imagen"}, status=400)

    foto = EvidencePhoto.objects.create(evidence=evidencia, image=image)

    return JsonResponse({
        "ok": True,
        "photo_url": foto.image.url
    })


# CHECKLIST

@require_POST
@login_required
def actualizar_checklist(request, booking_id):
    evidencia = get_object_or_404(Evidence, booking_id=booking_id)

    if not es_trabajador(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)

    campo = request.POST.get("campo")
    valor = request.POST.get("valor") == "true"

    if hasattr(evidencia, campo):
        setattr(evidencia, campo, valor)
        evidencia.save()
        return JsonResponse({"ok": True})

    return JsonResponse({"error": "Campo inválido"}, status=400)


# ADMIN — VER EVIDENCIA COMPLETA

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def evidencia_admin_view(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    evidencia = booking.evidence

    return render(request, "evidencia_admin.html", {
        "booking": booking,
        "evidencia": evidencia,
        "photos": evidencia.photos.all(),
    })


# PDF — FUNCIONANDO CON CLOUDINARY

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader

import requests
from io import BytesIO


@login_required
def evidencia_pdf(request, booking_id):
    if not (request.user.is_staff or es_trabajador(request.user)):
        return HttpResponse("No autorizado", status=403)

    booking = get_object_or_404(Booking, id=booking_id)
    evidencia = booking.evidence
    photos = evidencia.photos.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Evidencia_{booking.customer_name}.pdf"'

    styles = getSampleStyleSheet()
    story = []
    doc = SimpleDocTemplate(response, pagesize=letter)

    # Título
    story.append(Paragraph("<b>Reporte de Evidencia - HappyHuapi</b>", styles['Title']))
    story.append(Spacer(1, 12))

    # Info del evento
    story.append(Paragraph(f"<b>Cliente:</b> {booking.customer_name}", styles['Normal']))
    story.append(Paragraph(f"<b>Fecha:</b> {booking.event_date}", styles['Normal']))
    story.append(Paragraph(f"<b>Horario:</b> {booking.start_time} - {booking.end_time}", styles['Normal']))
    story.append(Paragraph(f"<b>Lugar:</b> {booking.location or '—'}", styles['Normal']))
    story.append(Spacer(1, 15))

    # Checklist
    check = lambda x: "✔" if x else "✘"
    story.append(Paragraph("<b>Checklist</b>", styles['Heading2']))
    story.append(Paragraph(f"Mesa instalada: {check(evidencia.table_installed)}", styles['Normal']))
    story.append(Paragraph(f"Decoración lista: {check(evidencia.decoration_ready)}", styles['Normal']))
    story.append(Paragraph(f"Globos inflados: {check(evidencia.balloons_ready)}", styles['Normal']))
    story.append(Paragraph(f"Candy listo: {check(evidencia.candy_ready)}", styles['Normal']))
    story.append(Spacer(1, 15))

    # Fotos
    story.append(Paragraph("<b>Fotos del montaje</b>", styles['Heading2']))
    story.append(Spacer(1, 10))

    for p in photos:
        try:
            r = requests.get(p.image.url)
            if r.status_code == 200:
                img_data = BytesIO(r.content)
                img = Image(img_data, width=250, height=250)
                story.append(img)
                story.append(Spacer(1, 12))
        except:
            pass

    doc.build(story)
    return response


# FINALIZAR EVIDENCIA

@require_POST
@login_required
def finalizar_evidencia(request, booking_id):
    evidencia = get_object_or_404(Evidence, booking_id=booking_id)

    if not es_trabajador(request.user):
        return JsonResponse({"error": "No autorizado"}, status=403)

    evidencia.completed = True
    evidencia.completed_at = timezone.now()
    evidencia.save()

    pdf_url = reverse("evidencia_pdf", args=[booking_id])

    return JsonResponse({"ok": True, "pdf_url": pdf_url})
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from catalog.models import Product
from .models import Booking

from django.http import JsonResponse

@login_required
@user_passes_test(is_admin)
def eliminar_producto(request, booking_id, product_id):
    try:
        reserva = Booking.objects.get(pk=booking_id)
        reserva.products.remove(product_id)
        reserva.save()
        return JsonResponse({"ok": True})
    except:
        return JsonResponse({"ok": False})
    from django.http import JsonResponse
from catalog.models import Product
from .models import Booking

#   ELIMINAR PRODUCTO DE UNA RESERVA
@login_required
@user_passes_test(is_admin)
def eliminar_producto_reserva(request, booking_id, product_id):
    try:
        reserva = Booking.objects.get(pk=booking_id)
        producto = Product.objects.get(pk=product_id)

        reserva.products.remove(producto)

        return JsonResponse({"ok": True})

    except Booking.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Reserva no existe"})

    except Product.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Producto no existe"})

@login_required
@user_passes_test(is_admin)
def agregar_producto(request, booking_id, product_id):
    """
    Agrega un producto a la reserva (uso desde AJAX en el modal).
    NO borra nada, solo suma.
    """
    try:
        reserva = Booking.objects.get(pk=booking_id)
        producto = Product.objects.get(pk=product_id)

        reserva.products.add(producto)

        return JsonResponse({"ok": True, "product_id": product_id})

    except Booking.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Reserva no existe"})

    except Product.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Producto no existe"})
def sobre_nosotros(request):
    return render(request, "sobre_nosotros.html")
def contacto(request):
    return render(request, "contacto.html")

